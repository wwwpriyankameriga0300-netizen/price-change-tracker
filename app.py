import streamlit as st
from engine import generate_price_change_report
from io import BytesIO
from datetime import datetime

from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.formatting.rule import FormulaRule
from openpyxl.utils import get_column_letter


st.set_page_config(page_title="Price Change Tracker", layout="wide")
st.title("📊 Price Change Tracker")

st.info(
    "• Comparison is based on Listing ID\n"
    "• Only increased / decreased items are shown\n"
    "• Red = price drop, Green = price increase"
)

files = st.file_uploader(
    "Upload Excel Files",
    type=["xlsx"],
    accept_multiple_files=True
)

if st.button("🚀 Generate Price Change Report"):

    if not files or len(files) < 2:
        st.error("Please upload at least TWO Excel files")
    else:
        with st.spinner("Processing files, please wait..."):
            df = generate_price_change_report(files)

        if df.empty:
            st.warning("No price changes detected.")
        else:
            st.success("✅ Price change report generated")
            st.dataframe(df, use_container_width=True)

            # ---------- WRITE EXCEL ----------
            temp_output = BytesIO()
            df.to_excel(temp_output, index=False)
            temp_output.seek(0)

            wb = load_workbook(temp_output)
            ws = wb.active

            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            # Find Change_Amount column
            change_col_letter = None
            for col in ws.iter_cols(1, ws.max_column):
                if col[0].value == "Change_Amount":
                    change_col_letter = col[0].column_letter
                    break

            last_col_letter = get_column_letter(ws.max_column)

            if change_col_letter:
                # 🟢 GREEN → price increase
                ws.conditional_formatting.add(
                    f"A2:{last_col_letter}{ws.max_row}",
                    FormulaRule(
                        formula=[f"VALUE(${change_col_letter}2)>0"],
                        fill=green_fill
                    )
                )

                # 🔴 RED → price decrease
                ws.conditional_formatting.add(
                    f"A2:{last_col_letter}{ws.max_row}",
                    FormulaRule(
                        formula=[f"VALUE(${change_col_letter}2)<0"],
                        fill=red_fill
                    )
                )

            final_output = BytesIO()
            wb.save(final_output)
            final_output.seek(0)

            st.download_button(
                "⬇️ Download Excel",
                final_output,
                file_name=f"price_change_report_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
